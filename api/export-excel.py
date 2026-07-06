from http.server import BaseHTTPRequestHandler
import json, os
from urllib.parse import urlparse, parse_qs
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ── Colores exactos del formato JL Advisory ─────────────────────────
C_NAVY,C_BLUE,C_SALMON,C_ORANGE,C_ORANGE2 = "FF1F3864","FF2E5FA3","FFFCE4D6","FF843C0C","FF833C00"
C_YELLOW,C_YELLOW2,C_LGRAY,C_WHITE = "FFFFF2CC","FF7F6000","FFF2F2F2","FFFFFFFF"
C_FOOTER,C_FOOTER2 = "FFD6E4F0","FF1F3864"

def solid(h): return PatternFill("solid", fgColor=h)
def thin(): return Side(style='thin', color="FFD1D5DB")
def med():  return Side(style='medium', color="FF2E5FA3")
def bord(top_med=False):
    t = med() if top_med else thin()
    return Border(left=thin(), right=thin(), top=t, bottom=thin())
def fnt(bold=False, size=9, color="FF000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def generate_excel(nombre: str, today: str, items: list) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Solicitud de Información"
    for i, w in enumerate([6.13,28.0,50.75,14.0,17.5,26.25,21.0], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fila 1: Título
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = f"{nombre}  —  Solicitud de Información  |  Due Diligence"
    c.font = fnt(True,14,C_WHITE); c.fill = solid(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Fila 2: Subtítulo
    ws.row_dimensions[2].height = 21.75
    ws.merge_cells("A2:D2"); ws.merge_cells("E2:G2")
    for addr, txt in [("A2",f"Documento de uso externo — para envío al vendedor  |  Emitido: {today}"),("E2","Confidencial — Uso exclusivo de las partes")]:
        c = ws[addr]; c.value = txt
        c.font = fnt(False,9,C_WHITE); c.fill = solid(C_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Fila 3: Leyenda
    ws.row_dimensions[3].height = 19.5
    for col,val,fc,fcolor,bold in [("A","●",C_SALMON,C_ORANGE,True),("B","PENDIENTE — aún no enviado",C_SALMON,C_ORANGE,True),("C","● INCOMPLETO — enviado pero faltan elementos",C_SALMON,C_ORANGE2,True),("D","Completar col. E y F",C_YELLOW,C_YELLOW2,True)]:
        c = ws[f"{col}3"]; c.value = val
        c.font = fnt(bold,9,fcolor); c.fill = solid(fc)
        c.alignment = Alignment(horizontal="center" if col=="A" else "left", vertical="center")
    ws.merge_cells("E3:G3")
    c = ws["E3"]; c.value = "Columna E: fecha estimada de envío  |  Columna F: observaciones del vendedor  |  Columna G: compromiso de entrega previo a la seña (según el vendedor)"
    c.font = fnt(False,9,C_YELLOW2); c.fill = solid(C_YELLOW)
    c.alignment = Alignment(horizontal="left", vertical="center", wrapText=True)

    # Fila 4: Headers
    ws.row_dimensions[4].height = 30
    for i,h in enumerate(["N°","Documento / Ítem requerido","Qué necesitamos exactamente y cómo enviarlo","Estado","Fecha comprometida\nde envío","Observaciones del vendedor\n(completar aquí)","Entrega Previo a Seña\n(según el vendedor)"],1):
        c = ws.cell(row=4,column=i,value=h)
        c.font = fnt(True,9,C_WHITE); c.fill = solid(C_NAVY)
        c.alignment = Alignment(horizontal="center",vertical="center",wrapText=True); c.border = bord()

    # Datos
    cur, ridx, secs = 5, 0, set()
    for it in items:
        sec = it.get("seccion","")
        if sec not in secs:
            secs.add(sec); ws.row_dimensions[cur].height = 21.75
            ws.merge_cells(f"A{cur}:G{cur}")
            c = ws.cell(row=cur,column=1,value=sec)
            c.font = fnt(True,10,C_WHITE); c.fill = solid(C_BLUE)
            c.alignment = Alignment(horizontal="left",vertical="center"); c.border = bord(True)
            cur += 1

        partes = []
        if it.get("cobertura"): partes.append(f"Se recibió: {it['cobertura']}")
        if it.get("faltantes"): partes.append(f"\nFalta: {it['faltantes']}")
        if it.get("como_cumplimentar"): partes.append(f"\n{it['como_cumplimentar']}")
        if it.get("alertas"): partes.append(f"\n⚠ {it['alertas']}")
        que_nec = "\n".join(partes).strip()
        estado = "Incompleto" if it.get("estado")=="Parcial" else (it.get("estado") or "Pendiente")
        antes = "SÍ (Información Básica/Estructural)" if it.get("antes_sena") else "NO (Reservar para Post-Seña/Contrato)"
        obs = "\n".join(l for l in (it.get("notas") or "").split("\n") if l.strip() and "Due Diligence (IA" not in l and "(3/7/2026 —" not in l)
        bg = C_LGRAY if ridx%2==0 else C_WHITE
        ws.row_dimensions[cur].height = 117

        for col,val,fc,align,bold,sz,fcolor in [(1,it.get("n_item",""),bg,"center",True,10,"FF000000"),(2,it.get("documento",""),bg,"left",True,10,"FF000000"),(3,que_nec,bg,"left",False,9,"FF000000"),(4,estado,C_SALMON,"center",True,10,C_ORANGE2 if estado=="Incompleto" else C_ORANGE),(5,"",C_YELLOW,"center",False,9,C_YELLOW2),(6,obs,C_YELLOW,"left",False,9,C_YELLOW2),(7,antes,C_YELLOW,"center",antes.startswith("SÍ"),9,C_YELLOW2)]:
            c = ws.cell(row=cur,column=col,value=val)
            c.font = fnt(bold,sz,fcolor); c.fill = solid(fc)
            c.alignment = Alignment(horizontal=align,vertical="top",wrapText=True); c.border = bord()
        cur+=1; ridx+=1

    # Pie
    ws.row_dimensions[cur].height = 19.5; ws.merge_cells(f"A{cur}:G{cur}")
    c = ws.cell(row=cur,column=1,value="Para consultas sobre esta solicitud comunicarse con el equipo de due diligence. Las columnas en amarillo son para completar por la empresa.")
    c.font = fnt(False,9,C_FOOTER2,True); c.fill = solid(C_FOOTER)
    c.alignment = Alignment(horizontal="left",vertical="center")

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)
            case_id = params.get('caseId',[''])[0]
            modo    = params.get('modo',['vendedor'])[0]
            
            if not case_id:
                self.send_response(400); self.end_headers()
                self.wfile.write(b'{"error":"Falta caseId"}'); return

            # Leer datos de Supabase
            from supabase import create_client
            sb = create_client(os.environ['SUPABASE_URL'], os.environ['SUPABASE_SERVICE_ROLE_KEY'])
            
            case_res = sb.table('dd_cases').select('nombre').eq('id', case_id).single().execute()
            reqs_res = sb.table('dd_case_requirements').select('*').eq('case_id', case_id).order('seccion_orden').order('n_item').execute()
            
            nombre = case_res.data.get('nombre','Due Diligence')
            from datetime import date
            today = date.today().strftime('%-d/%-m/%Y')
            all_items = reqs_res.data or []
            items = all_items if modo == 'interno' else [r for r in all_items if r.get('estado') != 'Recibido']
            
            xlsx_bytes = generate_excel(nombre, today, items)
            safe = nombre.replace(' ','_')[:30]
            fname = f"Solicitud_{safe}_{today.replace('/','')}.xlsx"
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.end_headers()
            self.wfile.write(xlsx_bytes)

        except Exception as e:
            self.send_response(500); self.send_header('Content-Type','application/json'); self.end_headers()
            self.wfile.write(json.dumps({'error': str(e)}).encode())

    def log_message(self, format, *args): pass
